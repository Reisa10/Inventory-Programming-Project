import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

Inventory_Database = "Database.xlsx"
Sales_Database = "sale.xlsx"
user_database = "user.xlsx"
Movement_Database = "inventory_movements.xlsx"

def create_if_not_exists(filename, headers):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(filename)

create_if_not_exists(Inventory_Database, ["product_id","product_name","category","price","stock_quantity","reorder_level"])
create_if_not_exists(Sales_Database, ["sale_id","date","product_id","quantity","unit_price","total"])
create_if_not_exists(user_database, ["username","password","role"])
create_if_not_exists(Movement_Database, ["movement_id","product_id","movement_type","quantity","date","remarks"])

wb1 = load_workbook(Inventory_Database)
ws1 = wb1.active
wb2 = load_workbook(Sales_Database)
ws2 = wb2.active
wb3 = load_workbook(user_database)
ws3 = wb3.active
wb4 = load_workbook(Movement_Database)
ws4 = wb4.active

def safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0

def safe_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0

# Safe input functions
def input_int(prompt):
    while True:
        value = input(prompt)
        try:
            value_int = int(value)
            if value_int < 0:
                print("Value cannot be negative. Try again.")
                continue
            return value_int
        except ValueError:
            print("Invalid input. Please enter a number.")

def input_float(prompt):
    while True:
        value = input(prompt)
        try:
            value_float = float(value)
            if value_float < 0:
                print("Value cannot be negative. Try again.")
                continue
            return value_float
        except ValueError:
            print("Invalid input. Please enter a number.")

def log_movement(product_id, movement_type, quantity, remarks):
    # Ensure product_id and movement_type are valid
    if not product_id:
        product_id = "-"
    if not movement_type:
        movement_type = "-"

    # Ensure quantity is integer
    quantity = safe_int(quantity)

    # Generate unique movement_id
    # Starts at 1 if only headers exist
    movement_id = ws4.max_row
    if movement_id < 1:
        movement_id = 1
    else:
        movement_id += 1

    date_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws4.append([movement_id, product_id, movement_type, quantity, date_now, remarks])
    try:
        wb4.save(Movement_Database)
    except PermissionError:
        print("Error: Close the inventory_movements.xlsx file before logging movement.")

def generate_sale_id():
    # Ensure unique sale ID
    sale_id = ws2.max_row
    if sale_id < 1:
        sale_id = 1
    else:
        sale_id += 1
    return sale_id

def add_new(product_Id, name, category, price, qty, rodlvl):
    ws1.append([product_Id.upper(), name.title(), category.title(), safe_float(price), safe_int(qty), safe_int(rodlvl)])
    wb1.save(Inventory_Database)
    log_movement(product_Id.upper(), "IN", qty, "Initial stock")

def remove_product():
    pid = input("Enter Product ID to remove: ").strip().upper()
    for row in ws1.iter_rows(min_row=2):
        if str(row[0].value).strip().upper() == pid:
            name = row[1].value if row[1].value else "-"
            while True:
                confirm = input(f"Are you sure you want to remove {name}? (y/n): ").strip().lower()
                if confirm == "y":
                    ws1.delete_rows(row[0].row, 1)
                    wb1.save(Inventory_Database)
                    log_movement(pid, "OUT", 0, "Product removed")
                    print(f"Product {pid}, {name} removed successfully")
                    return
                elif confirm == "n":
                    print("Removal cancelled")
                    return
                else:
                    print("Invalid input. Please enter 'y' or 'n'.")
    print("Product not found.")

def change_stock(product_id, new_stock):
    for row in ws1.iter_rows(min_row=2):
        if str(row[0].value).strip().upper() == product_id.upper():
            old_stock = safe_int(row[4].value)
            diff = safe_int(new_stock) - old_stock
            row[4].value = safe_int(new_stock)
            wb1.save(Inventory_Database)
            movement_type = "IN" if diff > 0 else "OUT"
            log_movement(product_id, movement_type, abs(diff), "Manual stock change")
            return True
    return False

def get_price(product_id):
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().upper() == product_id.upper():
            return safe_float(row[3])
    return None

def update_stock(product_id, qty_sold):
    for row in ws1.iter_rows(min_row=2):
        if str(row[0].value).strip().upper() == product_id.upper():
            current_stock = safe_int(row[4].value)
            new_stock = current_stock - safe_int(qty_sold)
            if new_stock < 0:
                print(f"Not enough stock for {product_name(product_id)}! Only {current_stock} left.")
                return False
            row[4].value = new_stock
            wb1.save(Inventory_Database)
            log_movement(product_id, "OUT", qty_sold, "Stock adjustment")
            return True
    return False

def buy(product_id, qty, sale_id):
    if not update_stock(product_id, qty):
        print("Sale failed. Stock not updated.")
        return False
    price = get_price(product_id)
    subtotal = safe_float(price) * safe_int(qty)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2.append([sale_id, now, product_id, qty, price, subtotal])
    wb2.save(Sales_Database)
    log_movement(product_id, "SALE", qty, "Customer purchase")
    return True

def get_pid_by_name(name):
    name = name.strip().title()  # match formatting in Excel
    for row in ws1.iter_rows(min_row=2, values_only=True):
        pid, pname = row[0], row[1]
        if pname and pname.strip().title() == name:
            return pid
    return None

def product_name(product_id):
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().upper() == product_id.upper():
            return row[1] or "-"
    return "-"

def print_receipt(sale_id):
    items = []
    sale_date = None

    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] == safe_int(sale_id):
            sale_date = row[1]
            pid = row[2]
            items.append({"pid": pid,"name": product_name(pid),"qty": safe_int(row[3]),"price": safe_float(row[4]),"total": safe_float(row[5])})
    if not items:
        return None
    return {"sale_id": safe_int(sale_id),"date": sale_date,"items": items}



def save():
    wb1.save("Database.xlsx")
    wb2.save("sale.xlsx")
    wb3.save("user.xlsx")
    wb4.save("inventory_movements.xlsx")

def list_products():
    print("\n====== INVENTORY LIST ======")
    print(f"{'ID':<10} {'Name':<20} {'Category':<15} {'Price':<10} {'Stock':<10} {'Reorder':<10}")
    
    for row in ws1.iter_rows(min_row=2, values_only=True):
        pid, name, cat, price, stock, reorder = row
        print(f"{pid or '-':<10} {name or '-':<20} {cat or '-':<15} {safe_float(price):<10} {safe_int(stock):<10} {safe_int(reorder):<10}")

def list_sales():
    print("\n====== SALES RECORDS ======")
    print(f"{'Sale ID':<10} {'Date':<20} {'Product ID':<10} {'Qty':<5} {'Unit Price':<10} {'Total':<10}")
    print("-"*70)
    for row in ws2.iter_rows(min_row=2, values_only=True):
        sid, date, pid, qty, price, total = row
        print(f"{sid or '-':<10} {date or '-':<20} {pid or '-':<10} {safe_int(qty):<5} {safe_float(price):<10} {safe_float(total):<10}")

def list_inventory_movements():
    print("\n====== INVENTORY MOVEMENTS ======")
    print(f"{'ID':<5} {'Product ID':<10} {'Type':<10} {'Qty':<5} {'Date':<20} {'Remarks':<20}")
    print("-"*75)
    for row in ws4.iter_rows(min_row=2, values_only=True):
        mid, pid, mtype, qty, date, remarks = row
        print(f"{mid or '-':<5} {pid or '-':<10} {mtype or '-':<10} {safe_int(qty):<5} {date or '-':<20} {remarks or '-':<20}")

def sales_summary():
    total_sales = total_items = total_revenue = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        total_sales += 1 if row[0] is not None else 0
        total_items += safe_int(row[3])
        total_revenue += safe_float(row[5])
    print("\n====== SALES SUMMARY ======")
    print(f"Total Sales Transactions: {total_sales}")
    print(f"Total Items Sold       : {total_items}")
    print(f"Total Revenue          : {total_revenue}")

def best_selling_products():
    product_sales = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            pid = row[2]
            if pid is None:
                continue
            qty = safe_int(row[3])
            product_sales[pid] = product_sales.get(pid, 0) + qty

    if not product_sales:
        print("No sales recorded yet.")
        return

    sorted_sales = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)

    print("\n====== BEST-SELLING PRODUCTS ======")
    print(f"{'Product ID':<10} {'Product Name':<20} {'Total Sold':<10}")
    print("-"*45)

    for pid, total in sorted_sales[:10]: 
        name = product_name(pid) or "-"
        pid = str(pid)
        total = total if total is not None else 0
        print(f"{pid:<10} {name:<20} {total:<10}")


def low_stock_alert():
    print("\n====== LOW STOCK ALERTS ======")
    print(f"{'Product ID':<10} {'Name':<20} {'Stock':<10} {'Reorder':<10}")
    low_stock_found = False

    low_stock_items = []

    for row in ws1.iter_rows(min_row=2, values_only=True):
        pid, name, cat, price, stock, reorder = row
        pid = str(pid).strip() if pid else "-"
        name = name.strip() if name else "-"
        stock = safe_int(stock)
        reorder = safe_int(reorder)

        if stock <= reorder:
            low_stock_items.append((pid, name, stock, reorder))

    if low_stock_items:
        low_stock_items.sort(key=lambda x: x[2])
        for pid, name, stock, reorder in low_stock_items:
            print(f"{pid:<10} {name:<20} {stock:<10} {reorder:<10}")
    else:
        print("All products have sufficient stock.")

# ----------------- TERMINAL MENU -----------------
def main_menu():
    while True:
        print("\n====== SALES & INVENTORY SYSTEM ======")
        print("1. Add New Product")
        print("2. Remove Product")
        print("3. Buy Product")
        print("4. Change Stock (Manual)")
        print("5. Check Product Price")
        print("6. Print Receipt")
        print("7. List All Products")
        print("8. List All Sales")
        print("9. List Inventory Movements")
        print("10. Sales Summary")
        print("11. Best-Selling Products")
        print("12. Low Stock Alerts")
        print("13. Exit")


        choice = input("Select option: ")

        if choice == "1":
            pid = input("Product ID: ").upper()
            name = input("Product Name: ").title()
            cat = input("Category: ").title()
            price = input_float("Price: ")
            qty = input_int("Quantity: ")
            reorder = input_int("Reorder Level: ")
            add_new(pid, name, cat, price, qty, reorder)
            print(f"{name}, {cat} with {qty} Quantity added successfully")

        elif choice == "2": 
            remove_product()

        elif choice == "3": 
            cart = []

            while True:
                list_products()
                product_name_input = input("\nEnter product name to add to cart (or type 'done' to finish): ").strip()
                if product_name_input.lower() == "done":
                    break

                pid = get_pid_by_name(product_name_input)
                if pid is None:
                    print(f"Product '{product_name_input}' not found.")
                    continue

                qty = input_int(f"Quantity for {product_name_input}: ")
                if qty <= 0:
                    print("Quantity must be greater than 0.")
                    continue

   
                current_stock = None
                for row in ws1.iter_rows(min_row=2, values_only=True):
                    if str(row[0]).strip().upper() == pid.upper():
                        current_stock = safe_int(row[4])
                        break
                if current_stock is None:
                    print("Error: Product not found in inventory.")
                    continue
                if qty > current_stock:
                    print(f"Not enough stock for {product_name_input}. Only {current_stock} left.")
                    continue

    
                for item in cart:
                    if item['pid'] == pid:
                        item['qty'] += qty
                        break
                else:
                    price = get_price(pid)
                    cart.append({'pid': pid, 'name': product_name_input, 'qty': qty, 'price': price})

                print(f"{qty} x {product_name_input} added to cart.")

            if not cart:
                print("Cart is empty. Nothing to checkout.")
                continue


            print("\n====== SHOPPING CART ======")
            total_amount = 0
            for item in cart:
                subtotal = item['qty'] * item['price']
                total_amount += subtotal
                print(f"{item['name']:<20} {item['qty']:<5} x {item['price']:<10} = {subtotal:<10}")
            print(f"Total Amount: {total_amount}")

            confirm_checkout = input("Proceed to checkout? (y/n): ").strip().lower()
            if confirm_checkout != "y":
                print("Checkout cancelled.")
                continue

            sale_id = generate_sale_id()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Process each cart item
            for item in cart:
                # Update stock first
                if not update_stock(item['pid'], item['qty']):
                    print(f"Failed to process {item['name']}. Not enough stock.")
                    continue

                # Calculate subtotal
                subtotal = item['qty'] * item['price']

                # Save to sales worksheet
                ws2.append([sale_id, now, item['pid'], item['qty'], item['price'], subtotal])

                # Log inventory movement
                log_movement(item['pid'], "SALE", item['qty'], "Customer purchase")

            # Save changes to Excel
            wb1.save(Inventory_Database)
            wb2.save(Sales_Database)

            print("\n====== RECEIPT ======")
            print(f"SALE ID: {sale_id}")
            print("-" * 40)
            for item in cart:
                subtotal = item['qty'] * item['price']
                print(f"{item['name']:<20} {item['qty']:<5} x {item['price']:<10} = {subtotal:<10}")
            print(f"Total Amount: {total_amount}")
            print("Purchase successful!")

        elif choice == "4":
            pid = input("Product ID: ").upper()
            new_stock = input_int("New Stock Quantity: ")  
            if change_stock(pid, new_stock):
                print(f"Stock for {product_name(pid)} updated to {new_stock}")
            else:
                print("Product not found.")

        elif choice == "5":
            pid = input("Product ID: ").strip().upper()
            if not pid:
                print("Product ID cannot be empty.")
                continue

            price = get_price(pid)
            if price is not None:
                print(f"Price for {product_name(pid)} (ID: {pid}): {price}")
            else:
                print(f"Product with ID '{pid}' not found.")

        elif choice == "6":
            while True:
                sale_id_input = input("Enter Sale ID: ")
                if not sale_id_input.isdigit():
                    print("Invalid Sale ID. Please enter a numeric value.")
                    continue
                sale_id = int(sale_id_input)
                if sale_id <= 0:
                    print("Sale ID must be greater than 0.")
                    continue
                break

            receipt = print_receipt(sale_id)
            if receipt:
                print("\n====== RECEIPT ======")
                print(f"Sale ID: {receipt['sale_id']}")
                print(f"Date   : {receipt['date']}")
                print("-" * 60)
                grand_total = 0
                print(f"{'Product ID':<10} {'Product':<10} {'Qty':<5} {'Unit Price':<10} {'Total':<10}")
                print("-"*50)
                for item in receipt['items']:
                    print(f"{item['pid']:<10} {item['name']:<10} {item['qty']:<5} {item['price']:<10} {item['total']:<10}")
                    grand_total += item['total']
                print("-"*50)
                print(f"{'TOTAL AMOUNT':<20} {'':<5} {'':<10} {grand_total:<10}")
            else:
                print("Sale not found.")

        elif choice == "7":
            list_products()
        elif choice == "8":
            list_sales()
        elif choice == "9":
            list_inventory_movements()
        elif choice == "10":
            sales_summary()
        elif choice == "11":
            best_selling_products()
        elif choice == "12":
            low_stock_alert()
        elif choice == "13":
            save()
            print("Exiting system...")
            break
        else:
            print("Invalid option")

main_menu()
